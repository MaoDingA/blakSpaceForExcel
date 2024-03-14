#!/usr/bin/env python
# coding=utf-8
import re
import openpyxl
from tkinter import filedialog, messagebox, Label, Button, Frame, Tk

class AppWindow(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.title = '霹雳无敌二字人名全角空格横通裂空融合升级工具'  # 更改窗口标题
        self.initUI()

    def initUI(self):
        self.master.title(self.title)
        self.master.geometry('400x150')  # 设置窗口大小

        # 文件选择标签和按钮
        self.file_label = Label(self.master, text='选择的文件: 无')
        self.file_label.pack()

        self.file_button = Button(self.master, text='选择Excel文件', command=self.openFileNameDialog)
        self.file_button.pack()

        # 开始处理按钮
        self.start_button = Button(self.master, text='开始处理', command=self.processWorkbook)
        self.start_button.pack()

    def openFileNameDialog(self):
        self.fileName = filedialog.askopenfilename(title="选择Excel文件", filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        if self.fileName:
            self.file_label.config(text=f'选择的文件: {self.fileName}')

    def processWorkbook(self):
        print("开始处理工作簿...")  # 日志
        if self.fileName:
            print(f"文件名: {self.fileName}")  # 日志
            result, changes = add_full_width_space_to_all_sheets(self.fileName)
            self.file_label.config(text='选择的文件: 无')
            result_message = result + '\n\n' + '\n'.join(changes) if changes else result
            messagebox.showinfo('处理结果', result_message)
        else:
            print("没有文件被选择")  # 日志

def add_full_width_space_to_all_sheets(filename):
    print(f"处理文件: {filename}")  # 打印文件名
    try:
        # 加载Excel工作簿
        workbook = openpyxl.load_workbook(filename)
        modified = False  # 添加一个标记来检查是否有单元格被修改
        changes = []  # 用于记录修改详情

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
                for cell in row:
                    if cell.value is not None:
                        # 删除所有空格、全角空格和制表符
                        original_value = cell.value
                        cleaned_value = str(cell.value).replace(' ', '').replace('　', '').replace('\t', '')

                        # 如果清理后的值是两个字符，添加全角空格
                        if len(cleaned_value) == 2:
                            cell.value = cleaned_value[0] + '　' + cleaned_value[1]  # 在两个字之间添加全角空格
                        else:
                            cell.value = cleaned_value

                        modified = True
                        changes.append(f"{sheet_name}!{cell.coordinate}: '{original_value}' -> '{cell.value}'")

        if modified:
            workbook.save(filename)  # 只有在至少有一个单元格被修改时才保存工作簿
            return '文件已修改并保存。', changes
        else:
            return '没有找到需要修改的单元格。', []
    except Exception as e:
        print(f"处理过程中发生错误: {e}")  # 打印错误信息
        return f"处理过程中发生错误: {e}", []

if __name__ == '__main__':
    root = Tk()
    app = AppWindow(root)
    app.mainloop()
